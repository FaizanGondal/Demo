import openpyxl


class HomePageData:

    test_HomePage_data =[{"FristName": "Rahul", "LastName": "Shetty", "Gender": "Male"},
                            {"FristName": "Aniska", "LastName": "Shetty", "Gender": "Female"}]
    @staticmethod
    def getTestData(testcase_name):
        dict= {}
        book = openpyxl.load_workbook("PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == testcase_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value


        return [dict]


